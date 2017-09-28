import xml.sax
import os, sys
import argparse
import openpyxl
from  openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side


class IprHandler_v2(xml.sax.ContentHandler):
    # Adapted from http://michaelrthon.com/runiprscan/

    # Tutorial on wtf is going on here: http://www.knowthytools.com/2010/03/sax-parsing-with-python.html
    # startElement is called called at every opening of a tag, endElement at the end.
    # If content (i.e. text between tags, not attributes within tags) are found then
    # self.characters() is called.
    # Only self.deets is used by get_IPRScan_xml_data() as this
    # class cannot be pickled easily so the data ends up as a list of dicts.
    def __init__(self):
        xml.sax.ContentHandler.__init__(self)
        self.deets = {'go num':set(), 'go term':set(), 'dom':set(), 'fam':set(),
                      'seq':''}
        self.seq = ''
    def startElement(self, name, attrs):

        #create a place for the sequence, overwriting any crap that's there so far
        if name == 'sequence':
            self.seq = ''

        if name == 'entry':
            # self.interpro.add(attrs['ac'] + ' ' + attrs['desc'])

            if attrs['type'] == "DOMAIN":
                # self.domain.add(attrs['desc'])
                self.deets['dom'].add(attrs['desc'])
                #print(attrs['desc'])

            if attrs['type'] == "FAMILY":
                # self.family.add(attrs['desc'])
                self.deets['fam'].add(attrs['desc'])

        if name == 'go-xref':
            # self.gonum.add(attrs['id'])
            # self.goterm.add(attrs['name'])
            self.deets['go term'].add(attrs['name'])
            self.deets['go num'].add(attrs['id'])


    def endElement(self, name):

        if name == 'sequence':
            self.deets['seq'] = [self.seq]
            #print(self.seq)
            self.seq = ''

    def characters(self, content):
        self.seq += content


def get_IPRScan_xml_data(dirname):
    # Adapted from http://michaelrthon.com/runiprscan/
    """returns a list of dicts containing filename and
    details specified in IprHandler_v2

    Keys should be:
        'filen' - name of the xml file
        'dom' - any domains
        'fam' - family membership of the protein
        'go num' & 'go term' - Gene Ontology numbers and terms.
    'filen' will be a string, the rest are set()"""

    all_deets = []

    file_list = os.listdir(dirname)

    for xmlfile in file_list:
        if not xmlfile.endswith('xml'):
            continue

        parser = xml.sax.make_parser()
        handler = IprHandler_v2()
        parser.setContentHandler(handler)
        parser.parse(open(dirname+'/'+xmlfile))
        #pdb.set_trace()
        deets = handler.deets
        deets['filen'] = xmlfile
        all_deets.append(deets)
        #return

    return all_deets


def make_excel_sheet(dirname,
                     save_filename = None,
                     sheet = None,
                     overwrite = False,
                     additional_cols = None,
                     add_cols_order = None,
                     deets_set = None):
    """Make an openpyxl.Workbook() containing details of results of
    a directory of IPRS xml results. The dir can contain non-xml
    files. Pass a save_filename and an Excel file will be created.
    Pass an openpyxl worksheet and the updated sheet will be returned.
    Pass neither and a freshly created Workbook will be returned.
    
    addictional_cols should be a dict with column headers as keys, and
    values either being a list of row values that should be ordered to
    match the IPRS results XML, or a dict with keys that match IPRS XML
    file names. add_cols_order is an optional list of key values in the
    order you want them on the final Excel sheet.

    Use deets_set if you're running get_IPRScan_xml_data seperately.
     """


    # Get list of dicts of relevant protein information
    dirname = os.path.abspath(dirname)
    assert os.path.isdir(dirname)

    # determine save file name.
    if save_filename:
        if  save_filename[-4:] not in ('xlsx', '.xls'):
            save_filename += '.xlsx'
        save_filename = os.path.abspath(save_filename)
        if os.path.isfile(save_filename) and not overwrite:
            print('File', save_filename,
                  'already exists.')
            cancel = input('Press enter to overwrite or type anything then enter to cancel.')
            if cancel:
                print('Cancelling...')
                return 0


    if sheet is None:
        wb = openpyxl.Workbook()
        active_sheet = wb.get_active_sheet()
    else:
        active_sheet = sheet

    # Define colours, using a bunch so that there's less chance reordering the sheet makes it confusing
    # Got these RGBA from matplotlib.cm 'Pastel1' colour map
    colrs =[(0.98431372549019602, 0.70588235294117652, 0.68235294117647061, 1.0),
            (0.70196078431372544, 0.80392156862745101, 0.8901960784313725, 1.0),
            (0.80000000000000004, 0.92156862745098034, 0.77254901960784317, 1.0),
            (0.87058823529411766, 0.79607843137254897, 0.89411764705882357, 1.0),
            (0.99607843137254903, 0.85098039215686272, 0.65098039215686276, 1.0),
            (1.0, 1.0, 0.80000000000000004, 1.0),
            (0.89803921568627454, 0.84705882352941175, 0.74117647058823533, 1.0),
            (0.99215686274509807, 0.85490196078431369, 0.92549019607843142, 1.0),
            (0.94901960784313721, 0.94901960784313721, 0.94901960784313721, 1.0)]

    # Lose the opacity value
    colrs = [(r,g,b) for r,g,b,a in colrs]

    # get slightly darkened versions
    dcolrs = [(r*0.9, g*0.9, b*0.9) for r, g, b in colrs]

    # # Interleve the dark and light colors
    # colrs = [val for pair in zip(colrs, dcolrs) for val in pair]

    # Alternate all light then all dark
    colrs += dcolrs

    # openpyxl uses hex rgb strings, this function produces them
    # assuming 0-255 values for rgb
    def int_to_hex(i):
        val = []
        for n in i:
            n = int(n)
            val.append(
                hex(n)[-2:]
            )
        return '00'+''.join(val)
    # get rgb hex strings
    colrs = [int_to_hex([r*255, g*255, b*255]) for r, g, b in colrs]
    # PatternFill objects are used to fill in openpyxl
    colrs = [PatternFill(start_color=c, end_color=c, fill_type='solid') for c in colrs]

    # Define row border colour
    row_colr = int_to_hex((100, 100, 100))
    row_border = Border(bottom=Side(style='thin', color=row_colr))

    # Headers
    headers = ['File Name', 'GO numbers', 'GO terms', 'Domains', 'Enzyme families', 'Sequence']
    if additional_cols or add_cols_order:
        if add_cols_order:
            headers = headers+add_cols_order
        else:
            add_cols_order = list(additional_cols.keys())
            headers = headers+add_cols_order
    active_sheet.append(headers)

    # Add results
    if deets_set is None:
        deets_set = get_IPRScan_xml_data(dirname)

    for result_i, deets in enumerate( deets_set ):
        deets_keys = ['go num', 'go term', 'dom', 'fam', 'seq']

        if add_cols_order:
            deets_keys+=add_cols_order
        # put single items into lists, convert sets
        fields = []
        for k in deets_keys:
            v = deets[k]
            if type(v) is set:
                v = list(v)
            elif type(v) is not list:
                v = [v]
            fields.append(v)

        # Get the longest list length
        depth = sorted([len(x) for x in fields])[-1]

        for i in range(depth):
            row = [deets['filen']]
            for field in fields:
                try:
                    row.append(field[i])
                except IndexError:
                    row.append('')
            #print(row)
            active_sheet.append(row)

    # Do colours. Each sequence's results occupies some rows, we want all rows
    # associated with a sequence to be the same colour
    prev_prot = None
    for rowi, row in enumerate(active_sheet.rows):
        # Check if this row is for a new result set
        if row[0].value != prev_prot:
            prev_prot = row[0].value
            # cycle through colors
            current_colour = colrs[rowi%len(colrs)]
        for cell in row:
            cell.fill = current_colour
            cell.border = row_border

    # book.save(paff+'GO terms 18Feb.xlsx')

    if sheet:
        return active_sheet
    elif save_filename:

        print('saving', save_filename)
        wb.save(save_filename)
    else:
        return wb

def run_from_command_line():
    parser = argparse.ArgumentParser(
        description= """Produce an Excel spreadsheet listing domain, enzyme family predictions, 
        and associated Gene Ontology terms from a directory of InterProScan XML files.""",
    )
    parser.add_argument("dir_name", help = 'Directory holding the IPRScan XML results files.')
    parser.add_argument('excel_file', help = 'Name and path of Excel that will be created.')
    parser.add_argument('-o', '--overwrite', action = 'store_true', default = False,
                        help = 'If the specified Excel file already exists, overwrite it without warning.')
    args = parser.parse_args()
    #print(args.overwrite)
    make_excel_sheet(args.dir_name, args.excel_file, overwrite = args.overwrite)

if __name__ == '__main__':
    # paff = r'C:\Users\JT\Dropbox\PhD\Experiments\Bioinfo\strain C comparison/'.replace('\\', '/')
    # fn = iprs_deets_fn = paff+'iprsdeets.pickle'
    # add_cols = ['bsr strain C', 'bsr RUH', 'bsr ANC', 'bsr bohem type',
    #             'strain C fasta title', 'E1 fasta title']
    # import pickle
    # iprsdeets = pickle.load(open(iprs_deets_fn, 'rb'))
    # iprsdeets = sorted(iprsdeets, key = lambda x: x['bsr strain C'], reverse=True)
    # make_excel_sheet(paff,
    #                  paff + 'IPRS strain C V3',
    #                  add_cols_order=add_cols,
    #                  deets_set=iprsdeets)
    run_from_command_line()

